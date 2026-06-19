#!/usr/bin/env python3
"""xlsx 명세 ↔ 코드 정합성 자동 갱신 (2026-06-04~05 머지 PR 6건 반영).

반영 대상 PR:
- PR #5  RAG 챗봇 UI (POST /chat/messages 신규)
- PR #6  임신 안전 안내(#21) + REQ-DASH-002 예상값 워터마크
- PR #7  What-if 시뮬레이션 (REQ-DASH-003, GET /dashboard/egfr-simulation)
- PR #9  회원가입 이메일 인증 (REQ-AUTH-003, 차단 정책)
- PR #10 Gmail SMTP 발송 모드 + demo fallback
- PR #12 슬럼프 + 마이크로 챌린지 (REQ-CHAL-006, 5종)

추가로 stale 정정 (어제 검증):
- REQ-DATA-006 / REQ-AUTH-007 / REQ-DASH-001 — 이미 풀구현돼 있는데 명세 미반영

운영 메모:
- xlsx 셀 서식·색상 일부는 코드 갱신 시 보존되지 않을 수 있음. 사용자가 시각 확인 권장.
- 실행: `python3 scripts/sync_api_spec.py`
"""

from __future__ import annotations

import openpyxl

REQ_XLSX = "docs/요구사항_정의서_v1.0_CKD.xlsx"
API_XLSX = "docs/API명세서_v1.0_CKD.xlsx"

# 기능 요구사항 시트 — 컬럼: A=REQ-ID, I=비고, J=상태
REQ_UPDATES: dict[str, tuple[str, str]] = {
    "REQ-AUTH-003": ("완료", "PR #9 — 6자리 코드 + 24h TTL + 미인증 403 차단"),
    "REQ-AUTH-007": ("완료", "5회 30분 잠금 + 테스트 5건 (2026-05-28 풀구현, memory 검증)"),
    "REQ-DATA-006": ("완료", "마이그12 — 운동(고/중강도)·좌식·결혼·가족력 3종 풀구현"),
    "REQ-DASH-001": ("완료", "7개 위젯 풀구현 (EgfrGauge·Risk·Egg·EgfrSim·Heatmap·Radial·Weekly)"),
    "REQ-DASH-002": ("완료", "PR #6 — EgfrGauge·RiskGauge 우상단 배지 + 내부 옅은 워터마크"),
    "REQ-DASH-003": ("완료", "PR #7 — What-if 시뮬레이션, 0~7일 일수 stepper 환산"),
    "REQ-CHAL-006": ("완료", "PR #12 — 5일 미체크인 슬럼프 감지 + 마이크로 챌린지 5종"),
}

# 신규 REQ (없으면 추가) — 컬럼 순서 A~K
NEW_REQS: list[tuple[str, ...]] = [
    (
        "REQ-DASH-005",
        "대시보드",
        "면책",
        "임신 안전 안내",
        "임신 체크 시 대시보드 상단에 산부인과 상담 권고·자가 식이/수분 제한 경고를 노출한다.",
        "기능",
        "높음",
        "P0",
        "PR #6 medical-review 라벨",
        "완료",
        "풀스택",
    ),
    (
        "REQ-AUTH-013",
        "회원관리",
        "이메일 발송",
        "Gmail SMTP + demo fallback",
        "EMAIL_MODE=gmail 발송 옵션. 실제 메일 송신과 응답에 코드 노출(fallback)을 병행해 스팸 분류·지연을 대비한다.",
        "기능",
        "보통",
        "P1",
        "PR #10",
        "완료",
        "풀스택",
    ),
]

# API 명세서 신규 엔드포인트 — origin 시트, R3 헤더, R4~ 데이터
# 컬럼 (1-base): A=- B=Request URL C=Method D=Short Desc E~G=Header/Query H~J=Req K~M=Resp N=- O=Status Code P=담당자 Q=완료 상태
NEW_APIS: list[tuple[str, str, str, str, str, str]] = [
    (
        "/api/v1/auth/email-verification/request",
        "POST",
        "REQ-AUTH-003 이메일 인증 코드 재발송 (PR #9, 1h 3회)",
        "200/400/422/429",
        "풀스택",
        "완료",
    ),
    (
        "/api/v1/auth/email-verification/verify",
        "POST",
        "REQ-AUTH-003 이메일 인증 코드 검증 (PR #9, 10/분)",
        "200/400/404/429",
        "풀스택",
        "완료",
    ),
    (
        "/api/v1/chat/messages",
        "POST",
        "REQ-RAG RAG 챗봇 질문/응답 (PR #5, 10/분)",
        "200/422/500/504",
        "AI+풀스택",
        "완료",
    ),
    (
        "/api/v1/dashboard/egfr-simulation",
        "GET",
        "REQ-DASH-003 예상 eGFR What-if 시뮬레이션 (PR #7) — 명세 v0.8 'POST /simulations/run' 갱신",
        "200",
        "풀스택",
        "완료",
    ),
    (
        "/api/v1/challenges/slump-micro",
        "GET",
        "REQ-CHAL-006 슬럼프 + 오늘의 마이크로 챌린지 (PR #12)",
        "200/404",
        "풀스택",
        "완료",
    ),
    (
        "/api/v1/challenges/slump-micro/checkin",
        "POST",
        "REQ-CHAL-006 마이크로 챌린지 체크인 (PR #12, 일별 중복 차단)",
        "200/400/404",
        "풀스택",
        "완료",
    ),
]


def update_requirements() -> list[str]:
    wb = openpyxl.load_workbook(REQ_XLSX)
    ws = wb["기능 요구사항"]
    log: list[str] = []
    existing: set[str] = set()
    for r in range(2, ws.max_row + 1):
        req_id = ws.cell(row=r, column=1).value
        if not req_id:
            continue
        existing.add(req_id)
        if req_id in REQ_UPDATES:
            new_status, new_note = REQ_UPDATES[req_id]
            ws.cell(row=r, column=9).value = new_note  # 비고
            ws.cell(row=r, column=10).value = new_status  # 상태
            log.append(f"  ✏️  {req_id}: 상태={new_status} / 비고 갱신")
    next_row = ws.max_row + 1
    for row_vals in NEW_REQS:
        if row_vals[0] in existing:
            log.append(f"  ⏭️  {row_vals[0]}: 이미 존재 — 건너뜀")
            continue
        for ci, val in enumerate(row_vals, start=1):
            ws.cell(row=next_row, column=ci).value = val
        log.append(f"  ➕ {row_vals[0]} 신규 추가 (행 {next_row})")
        next_row += 1
    wb.save(REQ_XLSX)
    return log


def update_api_spec() -> list[str]:
    wb = openpyxl.load_workbook(API_XLSX)
    ws = wb["origin"]
    log: list[str] = []
    # 기존 URL 인덱스 (중복 추가 방지)
    existing_urls: set[str] = set()
    for r in range(4, ws.max_row + 1):
        url = ws.cell(row=r, column=2).value
        if url:
            existing_urls.add(str(url).strip())
    next_row = ws.max_row + 1
    for url, method, desc, status_code, owner, complete in NEW_APIS:
        if url in existing_urls:
            log.append(f"  ⏭️  {method} {url}: 이미 존재 — 건너뜀")
            continue
        ws.cell(row=next_row, column=2).value = url
        ws.cell(row=next_row, column=3).value = method
        ws.cell(row=next_row, column=4).value = desc
        ws.cell(row=next_row, column=15).value = status_code  # O열
        ws.cell(row=next_row, column=16).value = owner  # P열
        ws.cell(row=next_row, column=17).value = complete  # Q열
        log.append(f"  ➕ {method} {url} 신규 추가 (행 {next_row})")
        next_row += 1
    wb.save(API_XLSX)
    return log


def main() -> None:
    print("=== 요구사항 정의서 갱신 ===")
    for line in update_requirements():
        print(line)
    print("\n=== API 명세서 갱신 ===")
    for line in update_api_spec():
        print(line)
    print("\n✅ 갱신 완료. xlsx를 직접 열어 셀 서식·색상 한 번 시각 확인 권장.")


if __name__ == "__main__":
    main()
